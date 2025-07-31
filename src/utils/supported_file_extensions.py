import asyncio
from pathlib import Path
from loguru import logger
from fpdf import FPDF  # pip install fpdf
from openpyxl import load_workbook

# ---------------------- Extension Sets ---------------------- #
text_file_extensions = [".txt", ".md", ".csv", ".json"]
word_file_extensions = [".docx", ".doc"]
excel_file_extensions = [".xlsx", ".xls"]
presentation_file_extensions = [".pptx", ".ppt"]

SUPPORTED_FILE_EXTENSIONS = [
    ".pdf",
    *text_file_extensions,
    *word_file_extensions,
    *excel_file_extensions,
    *presentation_file_extensions,
]


def autofit_excel_columns(path):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2  # Add some padding
    wb.save(path)


# ---------------------- Main Entry ---------------------- #
async def convert_supported_file_extension_to_pdf(file_path: Path) -> Path:
    """
    Convert a supported file to PDF and return the resulting PDF path.
    If already a PDF, return the original path.
    """
    ext = file_path.suffix.lower()

    if ext == ".pdf":
        return file_path

    logger.info(f"Converting {file_path} (type: {ext}) to PDF")

    pdf_path = file_path.with_suffix(".pdf")

    if ext in text_file_extensions:
        await convert_text_to_pdf(file_path, pdf_path)
        return pdf_path
    if ext in word_file_extensions:
        await convert_office_to_pdf(file_path, pdf_path)
        return pdf_path
    if ext in excel_file_extensions:
        autofit_excel_columns(file_path)
        await convert_office_to_pdf(file_path, pdf_path)
        return pdf_path
    if ext in presentation_file_extensions:
        await convert_office_to_pdf(file_path, pdf_path)
        return pdf_path
    raise ValueError(f"Unsupported file type for conversion: {ext}")


# ---------------------- Text/Markdown/CSV/JSON to PDF ---------------------- #


def ensure_font(
    pdf: FPDF, font_path: str = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
):
    # Đăng ký font Unicode
    pdf.add_font("DejaVu", "", font_path, uni=True)
    pdf.set_font("DejaVu", "", 12)


async def convert_text_to_pdf(file_path: Path, pdf_path: Path) -> None:
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
    except Exception as e:
        logger.error(f"Failed to read {file_path}: {e}")
        raise

    pdf = FPDF()
    pdf.add_page()
    ensure_font(pdf)

    for line in content.splitlines():
        try:
            pdf.cell(0, 10, txt=line, ln=True)
        except Exception as e:
            logger.error(f"Lỗi dòng: {line} | {e}")

    try:
        pdf.output(str(pdf_path))
    except Exception as e:
        logger.error(f"Failed to write PDF to {pdf_path}: {e}")
        raise
    logger.info(f"Converted {file_path} to {pdf_path}")


libreoffice_lock = asyncio.Lock()


# ---------------------- Office File to PDF (using LibreOffice) ---------------------- #
async def convert_office_to_pdf(
    file_path: Path,
    pdf_path: Path,
    timeout: int = 60,
) -> None:
    pdf_dir = pdf_path.parent  # also fix as previously recommended
    cmd = [
        "libreoffice",
        "--headless",
        "--convert-to",
        "pdf",
        "--outdir",
        str(pdf_dir),
        str(file_path),
    ]
    logger.info(f"Running: {' '.join(cmd)}")

    async with libreoffice_lock:
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
        )
        try:
            stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
        except asyncio.TimeoutError:
            proc.kill()
            logger.error(f"Timeout during conversion of {file_path}")
            raise RuntimeError("LibreOffice conversion timed out")

        if proc.returncode != 0:
            logger.error(f"LibreOffice failed: {stderr.decode().strip()}")
            raise RuntimeError(
                f"LibreOffice failed (exit {proc.returncode}): {stderr.decode().strip()}"
            )
        if not pdf_path.exists():
            logger.error(f"Expected output not found: {pdf_path}")
            raise FileNotFoundError(f"Output PDF not found: {pdf_path}")

    logger.info(f"Converted {file_path} to {pdf_path}")


# ---------------------- Example Usage ---------------------- #
# import asyncio
# asyncio.run(convert_supported_file_extension_to_pdf(Path("yourfile.docx")))
